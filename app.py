# import csv
# import os
# import random
# import sys
# import pandas as pd
# import openpyxl

# SUPERVISOR_FILE = "./data/canbocoithi.xlsx"
# ROOM_FILE = "./data/phongthi.xlsx"
# LOBBY_DEVISION_FILE = "./result/danhsachgiamsat.xlsx"
# ROOM_DEVISION_FILE = "./result/danhsachphancong.xlsx"

# def csv_to_xlsx(csv_file_path, xlsx_file_path):
#     # Read the CSV data
#     data = pd.read_csv(csv_file_path)
    
#     # Write the data to an XLSX file
#     data.to_excel(xlsx_file_path, index=False)

# def move_array(arr, number_of_times=1, reverse=False):
#     for _ in range(number_of_times):
#         if reverse:
#             arr.insert(0, arr.pop())
#         else:
#             arr.append(arr.pop(0))

#     return arr

# def get_xlsx(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active

#     data = []
#     for row in sheet.iter_rows(values_only=True):
#         data.append(row)

#     return data

# def get_csv(file_path, has_header=True):
#     with open(file_path, encoding='utf-8') as file:
#         csv_reader = csv.reader(
#             file,
#             delimiter=",",
#         )
#         is_header = has_header
#         data = []

#         for row in csv_reader:
#             if is_header:
#                 is_header = False
#                 continue

#             data.append(row)

#         return data

# if __name__ == "__main__":
#     args = sys.argv[1:]
#     if not (os.path.exists(SUPERVISOR_FILE) and os.path.exists(ROOM_FILE)):
#         print("File not found")
#         exit(1)

#     if os.path.exists(LOBBY_DEVISION_FILE):
#         os.remove(LOBBY_DEVISION_FILE)

#     if os.path.exists(ROOM_DEVISION_FILE):
#         os.remove(ROOM_DEVISION_FILE)

#     # rooms = get_csv(ROOM_FILE)
#     rooms = get_xlsx(ROOM_FILE)
#     # supervisors = get_csv(SUPERVISOR_FILE, has_header=False)
#     supervisors = get_xlsx(SUPERVISOR_FILE, has_header=False)
    
#     first_supervisors = supervisors[: len(supervisors) // 2]
#     second_supervisors = supervisors[len(supervisors) // 2 :]
#     if len(first_supervisors) != len(second_supervisors):
#         first_supervisors.append(["", "", "", ""])

#     with (
#         open(
#             ROOM_DEVISION_FILE, mode="w", newline="", encoding="utf-8"
#         ) as room_devision_file,
#         open(
#             LOBBY_DEVISION_FILE, mode="w", newline="", encoding="utf-8"
#         ) as lobby_devision_file,
#     ):
#         room_writer = csv.writer(
#             room_devision_file,
#             delimiter=",",
#             quotechar="'",
#             quoting=csv.QUOTE_MINIMAL,
#         )
#         lobby_writer = csv.writer(
#             lobby_devision_file,
#             delimiter=",",
#             quotechar="'",
#             quoting=csv.QUOTE_MINIMAL,
#         )

#         room_count = 0
#         lobby_count = 0
#         room_writer.writerow(
#             ["STT", "Mã GV", "Họ và tên", "Giám thị 1", "Giám thị 2", "Phòng thi"]
#         )
#         lobby_writer.writerow(["STT", "Mã GV", "Họ và tên", "Phòng thi được giám sát"])

#         for idx in range(int(args[0]) if args else 0):
#             first_supervisors = move_array(first_supervisors, 2)
#             second_supervisors = move_array(second_supervisors)

#         lobby_supervisors = []
#         for idx, room in enumerate(rooms):
#             if first_supervisors[idx][0] == "":
#                 lobby_supervisors.append(second_supervisors[idx])
#                 continue

#             if second_supervisors[idx][0] == "":
#                 lobby_supervisors.append(first_supervisors[idx])
#                 continue

#             room_writer.writerows(
#                 [
#                     (
#                         room_count + 1,
#                         first_supervisors[idx][0],
#                         first_supervisors[idx][1],
#                         "x",
#                         "",
#                         room[0],
#                     ),
#                     (
#                         room_count + 2,
#                         second_supervisors[idx][0],
#                         second_supervisors[idx][1],
#                         "",
#                         "x",
#                         room[0],
#                     ),
#                 ]
#             )
#             room_count += 2

#         if 2 * len(rooms) <= len(supervisors):
#             for idx in range(len(rooms), len(supervisors) // 2):
#                 lobby_supervisors.append(first_supervisors[idx])
#                 lobby_supervisors.append(second_supervisors[idx])

#         for lobby_supervisor in lobby_supervisors:
#             lobby_count += 1
#             supervisor_room = random.sample(rooms, 2)
#             lobby_writer.writerow(
#                 [
#                     lobby_count,
#                     lobby_supervisor[0],
#                     lobby_supervisor[1],
#                     f"Từ {supervisor_room[0][0]} đến {supervisor_room[1][0]}",
#                 ]
#             )

#     # Convert the files
#     csv_to_xlsx(LOBBY_DEVISION_FILE, LOBBY_DEVISION_FILE.replace('.csv', '.xlsx'))
#     csv_to_xlsx(ROOM_DEVISION_FILE, ROOM_DEVISION_FILE.replace('.csv', '.xlsx'))

import os
import random
import sys
import openpyxl

SUPERVISOR_FILE = "./data/canbocoithi.xlsx"
ROOM_FILE = "./data/phongthi.xlsx"
LOBBY_DEVISION_FILE = "./result/danhsachgiamsat.xlsx"
ROOM_DEVISION_FILE = "./result/danhsachphancong.xlsx"

def move_array(arr, number_of_times=1, reverse=False):
    for _ in range(number_of_times):
        if reverse:
            arr.insert(0, arr.pop())
        else:
            arr.append(arr.pop(0))

    return arr

def get_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def write_xlsx(data, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in data:
        sheet.append(row)

    workbook.save(file_path)

if __name__ == "__main__":
    args = sys.argv[1:]
    if not (os.path.exists(SUPERVISOR_FILE) and os.path.exists(ROOM_FILE)):
        print("File not found")
        exit(1)

    rooms = get_xlsx(ROOM_FILE)
    supervisors = get_xlsx(SUPERVISOR_FILE)
    
    first_supervisors = supervisors[: len(supervisors) // 2]
    second_supervisors = supervisors[len(supervisors) // 2 :]
    if len(first_supervisors) != len(second_supervisors):
        first_supervisors.append(["", "", "", ""])

    room_data = [["STT", "Mã GV", "Họ và tên", "Giám thị 1", "Giám thị 2", "Phòng thi"]]
    lobby_data = [["STT", "Mã GV", "Họ và tên", "Phòng thi được giám sát"]]

    room_count = 0
    lobby_count = 0

    for idx in range(int(args[0]) if args else 0):
        first_supervisors = move_array(first_supervisors, 2)
        second_supervisors = move_array(second_supervisors)

    lobby_supervisors = []
    for idx, room in enumerate(rooms):
        if first_supervisors[idx][0] == "":
            lobby_supervisors.append(second_supervisors[idx])
            continue

        if second_supervisors[idx][0] == "":
            lobby_supervisors.append(first_supervisors[idx])
            continue

        room_data.extend(
            [
                [
                    room_count + 1,
                    first_supervisors[idx][0],
                    first_supervisors[idx][1],
                    "x",
                    "",
                    room[0],
                ],
                [
                    room_count + 2,
                    second_supervisors[idx][0],
                    second_supervisors[idx][1],
                    "",
                    "x",
                    room[0],
                ],
            ]
        )
        room_count += 2

    if 2 * len(rooms) <= len(supervisors):
        for idx in range(len(rooms), len(supervisors) // 2):
            lobby_supervisors.append(first_supervisors[idx])
            lobby_supervisors.append(second_supervisors[idx])

    for lobby_supervisor in lobby_supervisors:
        lobby_count += 1
        supervisor_room = random.sample(rooms, 2)
        lobby_data.append(
            [
                lobby_count,
                lobby_supervisor[0],
                lobby_supervisor[1],
                f"Từ {supervisor_room[0][0]} đến {supervisor_room[1][0]}",
            ]
        )

    write_xlsx(room_data, ROOM_DEVISION_FILE)
    write_xlsx(lobby_data, LOBBY_DEVISION_FILE)